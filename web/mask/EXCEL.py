import csv
import logging
import os.path
from itertools import islice

import openpyxl
import xlrd
import pandas as pd

from web.config import STATIC_PATH
from web.core.utils import get_current_time


class XlsExcel:
    def __init__(self, file):

        self.sheets = xlrd.open_workbook(file).sheets()
        self.data = []
        self.sheet = self.sheets[0]

    def sheet_xuanze(self, n=0):
        self.sheet = self.sheets[n]
        return self.sheet

    def datas(self):
        datas = []
        for i in range(0, self.sheet.nrows):
            row_data = self.sheet.row_values(rowx=i, start_colx=0, end_colx=None)
            datas.append(row_data)
        logging.info('数据读取完毕'.center(33, '-'))
        return datas

    def all_title(self):
        return self.sheet.row_values(rowx=0, start_colx=0, end_colx=None)

    def bufen_titles(self, lie_titiles=None):
        if not lie_titiles:
            return self.all_title()
        title = []
        for i in lie_titiles:
            if i in self.all_title():
                title.extend([i])
        return title

    def count(self):
        return self.sheet.nrows - 1

    def search_lies(self, lie_titles):
        row_datas = []
        for i in range(1, self.sheet.nrows):
            row_data = []
            for lie_title in lie_titles:
                if lie_title in self.all_title():
                    row_data.extend(
                        [self.sheet.row_values(rowx=i, start_colx=0, end_colx=None)[self.search_index(lie_title)]])
                    # [i[self.search_index(lie_title)]
            row_datas.append(row_data)
        return row_datas

    def search_index(self, lie_title):
        if lie_title in self.all_title():
            return self.all_title().index(lie_title)

    @staticmethod
    def do_mask(func, filepath):
        xlrd.xlsx.ensure_elementtree_imported(False, None)
        xlrd.xlsx.Element_has_iter = True
        workbook = xlrd.open_workbook(filepath)
        sheet_names = workbook.sheet_names()
        # 遍历所有工作表，处理每个工作表的内容
        result_list = []
        try:
            for sheet_name in sheet_names:
                # 打开当前工作表
                worksheet = workbook.sheet_by_name(sheet_name)
                # 遍历当前工作表的所有行和列
                # sheet_data = [cell.value for row in worksheet.iter_rows() for cell in row[1:] if cell.value]
                value_list = [
                    int(worksheet.cell_value(row, col)) for row in range(worksheet.nrows) for col in
                    range(worksheet.ncols) if isinstance(worksheet.cell_value(row, col), float)]
                processed_value = func(value_list)
                result_list.append(processed_value)
        except Exception as e:
            return False, e
        return True, result_list[0]


def Excel(file):
    file_suffix = str(file).split('.')[-1]
    if file_suffix == 'xls':
        return XlsExcel(file)
    if file_suffix == 'xlsx':
        return XlsxExcel(file)
    if file_suffix == 'csv':
        from web.mask.mask import mask
        return csv_mask(mask, file)


class XlsxExcel:
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file, data_only=True)  # read_only=True,
        # self.ws = self.wb[self.wb.sheetnames[0]]
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.sheets = self.wb.sheetnames
        self.data = []
        self.file = file

    def sheet_xuanze(self, n=0):
        self.ws = self.wb[self.wb.sheetnames[n]]
        return self.ws

    @property
    def datas(self):
        data = []
        for row in self.ws.values:
            data.append(list(row))
        logging.info('数据读取完毕'.center(33, '-'))
        self.data = data[:]
        return self.data

    @property
    def all_title(self):
        return self.data[0]

    def bufen_titles(self, lie_titiles=None):
        if not lie_titiles:
            return self.all_title
        title = []
        for i in lie_titiles:
            if i in self.all_title:
                title.extend([i])
        return title

    @property
    def count(self):
        return self.ws.max_row - 1

    def search_lies(self, lie_titles):
        row_datas = []
        for i in range(1, self.ws.max_row):
            row_data = []
            for lie_title in lie_titles:
                if lie_title in self.all_title:
                    row_data.extend(
                        [self.ws.row_values(rowx=i, start_colx=0, end_colx=None)[self.search_index(lie_title)]])
            row_datas.append(row_data)
        return row_datas

    def search_index(self, lie_title):
        if lie_title in self.all_title:
            return self.all_title.index(lie_title)

    def write_one(self):
        self.ws.cell(1, 1).value = ''

    def to_do(self, func, filename):
        datas = self.datas
        if len(datas) == 0:
            datas = self.datas[:]
        else:
            datas = self.datas[1:]
        try:
            for n in range(len(self.sheets)):
                datas = self.sheet_xuanze(n).value
                for i in range(self.sheet_xuanze(n).value):
                    for j in range(len(datas[i])):
                        if isinstance(datas[i][j], float):
                            datas[i][j] = int(datas[i][j])
                        data = func(str(datas[i][j]))
                        datas[i][j] = data
        except Exception as e:
            return False, e
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in datas:
            ws.append(row)
        tmp_dir = f'mask-{get_current_time()}.xlsx'
        wb.save(get_file_path(tmp_dir))
        return True, tmp_dir

    def do_mask(self, func):
        result_list = []
        # try:
        for sheet_name in self.sheets:
            worksheet = self.wb[sheet_name]

            sheet_data = [int(cell.value) if isinstance(cell.value, float) else str(cell.value) for row in
                          worksheet.iter_rows(min_row=2) for cell in row]
            logging.info('~~~~~~~~~~~~~~~~~~~~~~~~sheet_data')
            logging.info(len(sheet_data))
            if sheet_data:
                result = func(sheet_data)
                result_list.append(result)
            del worksheet
        # except Exception as e:
        #     return False, e
        return True, result_list[0]

    def yield_mask(self, batch_size=100):
        chunk_set = set()
        for sheet_name in self.sheets:
            logging.info('sheet_name')
            logging.info(sheet_name)
            worksheet = self.wb[sheet_name]
            # 优化此段代码。导致列表太大内存占用过高
            # current_chunk_list = [cell.value for row in worksheet.iter_rows(min_row=2) for cell in row if cell.value]
            # if len(current_chunk_list) == batch_size:
            #     yield current_chunk_list
            #     current_chunk_list = []
            # else:
            #     yield current_chunk_list
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.value:
                        chunk_set.add(cell.value)
                        if len(chunk_set) == batch_size:
                            yield chunk_set
                            chunk_set = set()
            if chunk_set:
                yield chunk_set
            del worksheet

    def func_yield_mask(self, func):
        res_list = []
        try:
            for datas in self.yield_mask():
                sen_data = func(datas)
                res_list.extend(sen_data)
        except Exception as e:
            return False, e
        return True, res_list


def pandas_read_xlsx(path, func):
    excel_file = pd.ExcelFile(path)
    result_list = [{'文件路径': path}]
    # 获取所有工作表名称
    sheet_names = excel_file.sheet_names

    # 一次性读取所有工作表的内容
    excel_data = {sheet_name: excel_file.parse(sheet_name) for sheet_name in sheet_names}

    # 将所有工作表的数据合并为一个 DataFrame
    combined_data = pd.concat(excel_data.values(), ignore_index=True)
    logging.info('combined_data**********')
    logging.info(type(combined_data))
    pd_list = combined_data.to_dict('records')
    logging.info(pd_list)
    result = func(pd_list)
    result_list.append({'扫描结果': result})
    return True, result_list


def get_file_path(filename):
    return os.path.join(STATIC_PATH, filename)


def csv_mask(func, file):
    csv.field_size_limit(500 * 1024 * 1024)
    with open(file, 'r') as f:
        datas = [value for row in csv.reader(f) for value in row]
        logging.info('数据读取完毕'.center(33, '-'))
        res = func(datas)
    # except Exception as e:
    #     return False, e
    return True, res


def read_chunk_file(file_name, func, chunk_size=1024 * 8):
    data_list = []
    with open(file_name, 'r', encoding='utf-8') as file:
        while True:
            chunk = file.read(chunk_size)
            if not chunk:
                break
            # 处理每一块数据
            result = func(chunk)
            data_list.append(result)
    return data_list


def read_line_file(file_name, batch_size=1024):
    with open(file_name, 'r', encoding='utf-8') as file:
        while True:
            batch = list(islice(file, batch_size))
            if not batch:
                break
            yield batch


def func_yield_txt_mask(file_name, func):
    res_list = []
    try:
        for datas in read_line_file(file_name):
            sen_data = func(datas)
            res_list.append(sen_data)

    except Exception as e:
        return False, 'TXT文件不是utf-8编码处理错误，请修改后重新扫描！'

    return True, res_list[0]


def txt_mask(func, file):
    status, processed_line = func_yield_txt_mask(file_name=file, func=func)
    return status, processed_line
